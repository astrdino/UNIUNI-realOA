#The file the framework FLASK used to deliver Web App content
from flask import Flask, request,render_template,send_file,send_from_directory, jsonify

import pandas as pd
import numpy as np
import openpyxl 

import os
import io

import csv
from datetime import datetime, timedelta
import pytz
#import win32clipboard as win32CPB #Windows Clipboard

import threading
import time



#Version 01/24/2024

#https://realpython.com/python-web-applications/


#Create Flask instance "app"
app = Flask(__name__)

#Create directory for downloadable files
app.config['UPLOAD_FOLDER'] = 'tmp'
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

celsius = "N/A"
progress = 0

# @app.route("/Convert.html?celsius=123")
def fahrenheit_from(celsius):
    """Convert Celsius to Fahrenheit degrees."""
    try:
       
        fahrenheit = float(celsius) * 9 / 5 + 32
        fahrenheit = round(fahrenheit, 3)  # Round to three decimal places
        
        print(str(fahrenheit)+"from function")
        return fahrenheit
        # return render_template('Convert.html', my_var = str(fahrenheit))  
    except ValueError:
        return "invalid input"

@app.route("/Convert.html", methods=['GET', 'POST'])
def Convert():
    fahrenheit = None
    if request.method == 'POST':
        print('hello')
        celsius = request.form.get("celsius", "")
        fahrenheit = fahrenheit_from(celsius)
        print(celsius)
        print(fahrenheit)
        return render_template('Convert.html', my_var = str(fahrenheit))

    return  render_template('Convert.html')


@app.route("/Uniuni-Concat.html", methods=['GET', 'POST'])
def Uniuni_Concat():

    current_time = datetime.now().strftime("%m-%d-%Y")

    if request.method == 'POST':

        batchNum = request.form.get('batchInput')
        driverIDs = request.form.getlist("driverInputs[]")

        #Remove syntax
        for ele in driverIDs:
            print(ele)
            ele.replace("'",'')

        
        return render_template('Uniuni-Concat.html', batchNum=batchNum, driverIDs=driverIDs, current_time = current_time)

    return render_template('Uniuni-Concat.html', current_time = current_time)

# /////////////////////////////////Auto Daily Report/////////////////////////////////////////////////////

DATE_LIST = [] #Including the date needed to generate daily report


'''
    TOTAL
    1. Well-defined data structure for the application to sorting "date","batch number","total state result in terms of list" accordingly
    2. Each element in the list is a dictionary variable
    3. e.g. [{'date': '01-07-2024', 'batch': ['PHSUB-202401051550', 'PHX-YE-20240105'], 'st_result': [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]},{},{}]
'''
TOTAL = [] 


PATTERN_Bucket = [195,199,200,202,203,207,211,213,216,218,220,228,230,231]
MARK_States = [202,211,231] #Alarm config
ALARM_Collection = []


@app.route('/Auto_DailyReport.html', methods=['GET','POST'])
def upload_RA():

    #Upload Road Assignment File

    if 'file_roadAssignment' not in request.files:
        return render_template('Auto_DailyReport.html',Sheet_Status = 'Sheet not foud',Initial = 'Y')
    

    file_roadAssignment = request.files['file_roadAssignment']
    # file_orderList = request.files['file_orderList']
    

    if file_roadAssignment.filename == '':
        return 'No selected file'
    
    if file_roadAssignment: 
        # return render_template('Auto_DailyReport.html',RA_Sheet_Status = 'Road Assignment Found',output_filename=process_RA(file_roadAssignment))
        
        DATE_LIST = getDate()

        global TOTAL
        TOTAL = genMapList(file_roadAssignment,DATE_LIST)


        return render_template('Auto_DailyReport.html',RA_Sheet_Status = 'Road Assignment Found',batchNum=displayBatchNum(TOTAL) )
        # return render_template('Auto_DailyReport.html',RA_Sheet_Status = 'Road Assignment Found')

@app.route('/Auto_DailyReport_uploadOL.html', methods=['GET','POST'])
def upload_OL():

    #Upload Road Assignment File

    if 'file_orderList' not in request.files:
        return render_template('Auto_DailyReport.html',Sheet_Status = 'Sheet not foud')
    

    file_OL = request.files['file_orderList']
    # file_orderList = request.files['file_orderList']
    

    if file_OL.filename == '':
        return 'No selected file for Order List'
    
    if file_OL: 

        __TEST__start_Time = time.time()
        
        # print(TOTAL)
        print(countStates(file_OL,TOTAL))

        writeIn()

        """
        String Display On the Site
        """
        # Open the Excel file
        wb = openpyxl.load_workbook('./tmp/V3-Auto-Daily-Report.xlsx',data_only = True)
        sheet = wb.active


        # Convert the sheet to HTML
        output = io.StringIO()
        output.write('<table class="excel-table">')

        # Header row
        output.write('<tr>')
        for cell in sheet[1]:
            cell_value = '' if cell.value is None else cell.value
            output.write(f'<th>{cell_value}</th>')
        output.write('</tr>')

        # Data rows
        for row in sheet.iter_rows(min_row=2):
            output.write('<tr>')
            for cell in row:
                cell_value = '' if cell.value is None else cell.value

                #Make particular title bold
                if cell_value == 'Pakg Status' or cell_value == 'Quantity' or cell_value == "Total Rate" or cell_value == "TTL PAKGS" or is_date(str(cell_value)):
                    output.write(f'<td><strong>{cell_value}</strong></td>')
                else:
                    output.write(f'<td>{cell_value}</td>')
            output.write('</tr>')

        output.write('</table>')
        html_table = output.getvalue()
        output.close()

        __TEST__end_Time = time.time()

        print(__TEST__end_Time - __TEST__start_Time)

        return render_template('Auto_DailyReport.html',output_filename="V3-Auto-Daily-Report.xlsx", table=html_table,)
        # return render_template('Auto_DailyReport.html',RA_Sheet_Status = 'Road Assignment Found')


def getDate():

    # Define your local timezone
    local_timezone = pytz.timezone('America/Phoenix')  # e.g., 'America/New_York'

    # Get the current time in UTC
    utc_time = datetime.now(pytz.utc)

    # Convert UTC time to your local time
    local_time = utc_time.astimezone(local_timezone)

    Report_Date = local_time.strftime("%Y-%m-%d")

    #This is a string
    #Report_Date = datetime.now().strftime("%Y-%m-%d")
    #print(type(Report_Date))
    # Parse the date string into a datetime object
    Report_Date = datetime.strptime(Report_Date, '%Y-%m-%d')
   # print(type(Report_Date))

    

    # datemonth = Report_Date[0:2]
    # dateday= Report_Date[2:4]
    # dateyear= Report_Date[4:8]

    DateTracking_Start = 1
    DateTracking_End = 6


    # date-month-list
    dml = []
    # date-day-list
    ddl = []

    #Date List including maping list
    datelist = []
    

    for i in range(DateTracking_Start,DateTracking_End+1):

        date = Report_Date - timedelta(days=i)

        #Month Digit Formating
        if(date.month < 10):
            dml.append('0'+str(date.month))
        else:
            dml.append(str(date.month))

        #Date Digit Formating
        if(date.day < 10):
            ddl.append('0'+str(date.day))
        else:
            ddl.append(str(date.day))
    

        datelist = list(zip(dml,ddl))[::-1]

    #[('01', '06'), ('01', '07'), ('01', '08'), ('01', '09'), ('01', '10'), ('01', '11')]
    # print(datelist)
    return datelist

def is_date(string):
    try:
        datetime.strptime(string, '%m-%d-%Y')
        return True
    except ValueError:
        return False

def genMapList(file,DATE_LIST):

    
    #Read local excel to dictionary via pandas 
    DATA_DF = pd.read_excel(file,sheet_name=None)
    # print(DATA_DF.__len__())

    # print(DATA_DF)



    #Read-Write Subbatch number in each sheet
    """
    Format -> result = {} #Dictionary ['11/17': ('sub1,sub2',[STATECount Result])],..]

    """

    result = []
    
    for date in DATE_LIST:

        result_sub_ele = {'date':'','batch':[],'st_result':[]} #{'date': '11/03','batch':['PHX-YE-20231101','PHSUB-202311010828'], 'result':[0,0,...,0,0]}
    
        #
        ws_date_in = str(date[0])+'-'+str(date[1])
        single_date_result = [0] * len(PATTERN_Bucket)

        #Handle 2023->2024 Extention
        try:
            batch_tmp = str(DATA_DF[ws_date_in].at[1,'Unnamed: 2']).split(',')

        except KeyError:
            ws_date_in = ws_date_in + "-2024"
            batch_tmp = str(DATA_DF[ws_date_in].at[1,'Unnamed: 2']).split(',')

        # print(batch_tmp)
        
        result_sub_ele['date'] = ws_date_in
        result_sub_ele['batch'] = batch_tmp
        result_sub_ele['st_result'] = single_date_result
  

        result.append(result_sub_ele)
       


    # print(result)

    return result

def displayBatchNum(d_b_mapList):


    #Display string
    ds = ''

    for day in d_b_mapList:
        for sb in day['batch']:
            ds += sb + ','
    

    #Windows auto clipboard
    # win32CPB.OpenClipboard()
    # win32CPB.EmptyClipboard()
    # win32CPB.SetClipboardText(ds)
    # win32CPB.CloseClipboard()

    return ds

def countStates(file,emptyTotal):

    #Get downloaded order sheet
    sheet = openpyxl.load_workbook(file)['Order List']
    

    # Define the cell range 
    start_row = 2 #Column Name on 1st Row
    end_row = sheet.max_row
    state_column = 'C'
    batch_column = 'E'
    driver_column = 'H'


    #Check each row in the sheet
    for row in range(start_row,end_row + 1):


        state_value = sheet[f'{state_column}{row}'].value
        batch_value = sheet[f'{batch_column}{row}'].value

        res_state_index = PATTERN_Bucket.index(state_value)

        instance_date = '' #The date of the order in this row

        #print(emptyTotal)   
        
        for date in emptyTotal:
            
            #Find day by batch number
            res = [value for key,value in date.items() if any(batch == batch_value for batch in date['batch'])]
            
            #Update state_result for each day
            if(res != []):

                instance_date = res[0] #Record the date for 'this' row
                res[2][res_state_index]  += 1
                # print(res)



        #Sus states collector
        if(state_value in MARK_States):
        
            ALARM_Collection.append((instance_date,sheet[f'{driver_column}{row}'].value,state_value,batch_value))


        #Add Analysis
        #!! Not really "empty" now !!
        for ddata in emptyTotal: 
            

            #Total Packages in THIS day
            this_ttlPAKGs = sum(ddata['st_result'])

            ddata['data_analysis'] = []

            for one_st in ddata['st_result']:
                rate = 0
                if(this_ttlPAKGs != 0):
                    rate = one_st / this_ttlPAKGs

                ddata['data_analysis'].append('{:.2%}'.format(rate))
                "{:.0%}".format(1/3)


            ddata['data_analysis'].append(this_ttlPAKGs)
            
        #     # ddata['data_analysis'] = [this_ttlPAKGs]
                
        # print(emptyTotal)

    return emptyTotal

def writeIn():

    global progress

    #Write-in destination
    result_book = openpyxl.load_workbook('./UniuniHost/Daily_Report_Template.xlsx') #Get the Daily Report Template
    result_sheet = result_book['Master Form for report'] 

   

    """
    Write-in row&col range
    """

    #Regular
    result_row_first_3 = list(range(5,len(PATTERN_Bucket)+5))
    result_row_last_3 = list(range(24,len(PATTERN_Bucket)+24)) #End + 1

    result_col = ''

    dateWriteIn_row = 3
    dateWriteIn_col = ''

    ttlRate_col = '' #F, J, H
    

    #Alarm
    ALARMWriteIn_col_date = 'Q'
    ALARMWriteIn_col_driver = 'R'
    ALARMWriteIn_col_state = 'S'
    ALARMWriteIn_col_batNum = 'T'

    ALARMWriteIn_row_head = 5

    

    #Daily data iterating
    for cnt,val in enumerate(TOTAL):
 

        #Overwrite in 'Daily Report', otherwise loading 'Template'
        if(cnt > 0):
            result_book = openpyxl.load_workbook('./tmp/V3-Auto-Daily-Report.xlsx')
            result_sheet = result_book['Master Form for report'] 

        #End of last "if"

        if cnt == 0:
            result_col = 'E'
            dateWriteIn_col = 'D'
            dateWriteIn_row = 3

            ttlRate_col = 'F'
            
        elif cnt == 1:
            result_col = 'I'
            dateWriteIn_col = 'H'
            dateWriteIn_row = 3

            ttlRate_col = 'J'

        elif cnt == 2:
            result_col = 'M'
            dateWriteIn_col = 'L'
            dateWriteIn_row = 3

            ttlRate_col = 'N'

        elif cnt == 3:
            result_col = 'E'
            dateWriteIn_col = 'D'
            dateWriteIn_row = 22

            ttlRate_col = 'F'
            
        elif cnt == 4:
            result_col = 'I'
            dateWriteIn_col = 'H'
            dateWriteIn_row = 22

            ttlRate_col = 'J'

        elif cnt == 5:
            result_col = 'M'
            dateWriteIn_col = 'L'
            dateWriteIn_row = 22

            ttlRate_col = 'N'


        
        #Write in date
        result_sheet[dateWriteIn_col+str(dateWriteIn_row)] = val['date']
        result_book.save('./tmp/V3-Auto-Daily-Report.xlsx')

        


        #Write in state data
        for idx,st in enumerate(val['st_result']):
            #idx = # of each single state
            
            
            if(cnt < 3):
                #Write in "TTL PAKGS"
                result_sheet[result_col+str(result_row_first_3[-1] + 1)] =  val['data_analysis'][-1]
                result_sheet[ttlRate_col+str(result_row_first_3[-1] + 1)] =  "100.00%"

                #Write in "Quantity"
                result_sheet[result_col+str(result_row_first_3[idx])] = st 

                #Write in "Totol Rate"
                result_sheet[ttlRate_col+str(result_row_first_3[idx])] = val['data_analysis'][idx]
            else:
                result_sheet[result_col+str(result_row_last_3[-1] + 1)] = val['data_analysis'][-1] 
                result_sheet[ttlRate_col+str(result_row_last_3[-1] + 1)] =  "100.00%"
                result_sheet[result_col+str(result_row_last_3[idx])] = st
                result_sheet[ttlRate_col+str(result_row_last_3[idx])] = val['data_analysis'][idx]

            result_book.save('./tmp/V3-Auto-Daily-Report.xlsx')

        # #Write in data analysis
        # for idx,st in enumerate(val['data_analysis']):
        #     #idx = # of each single state
        #     if(cnt < 3):
        #         result_sheet[result_col+str(result_row_first_3[idx])] = st #data value
        #         result_sheet[ttlRate_col+str(result_row_first_3[idx])] = st #data analysis
        #     else:
        #         result_sheet[result_col+str(result_row_last_3[idx])] = st
        #         result_sheet[ttlRate_col+str(result_row_last_3[idx])] = st

        #     result_book.save('./tmp/V3-Auto-Daily-Report.xlsx')

        #Progress Management #For "Process Bar"
        progress = cnt + 1

        
        print(val['st_result'])


    #ALARM info write in
    for record in ALARM_Collection:

        result_book = openpyxl.load_workbook('./tmp/V3-Auto-Daily-Report.xlsx')
        result_sheet = result_book['Master Form for report'] 

        result_sheet[ALARMWriteIn_col_date+str(ALARMWriteIn_row_head)] = record[0]
        result_sheet[ALARMWriteIn_col_driver+str(ALARMWriteIn_row_head)] = record[1]
        result_sheet[ALARMWriteIn_col_state+str(ALARMWriteIn_row_head)] = record[2]
        result_sheet[ALARMWriteIn_col_batNum+str(ALARMWriteIn_row_head)] = record[3]

        result_book.save('./tmp/V3-Auto-Daily-Report.xlsx')

        ALARMWriteIn_row_head+=1




        

    return 1


#Download
@app.route('/download/<filename>')
def download_file(filename):
    print("last step")
    print(filename)
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

@app.route('/')
def index():
    # Define your local timezone
    local_timezone = pytz.timezone('America/Phoenix')  # e.g., 'America/New_York'

    # Get the current time in UTC
    utc_time = datetime.now(pytz.utc)

    # Convert UTC time to your local time
    local_time = utc_time.astimezone(local_timezone)

    current_time = local_time.strftime("%Y-%m-%d, %H:%M:%S")

    return render_template('index.html', current_time = current_time)


if __name__ == "__main__":
    app.run(host="127.0.0.1", port=8080, debug=True)










